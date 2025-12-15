# ! /usr/bin/python

import base64
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl

try:
    import oci
except ImportError:
    os.system("pip install oci")
    import oci

def get_instance(Target_Instance):
    config = oci.config.from_file()
    print("Config file", config)
    # Initialize service client with default config file
    secrets_client = oci.secrets.SecretsClient(config)
    get_secret_bundle_by_name_response = secrets_client.get_secret_bundle_by_name(
        secret_name=Target_Instance,
        vault_id="ocid1.vault.oc1.iad.ejurjoqaaahte.abuwcljsmbmhq7elzyjpfdweajchv3nvjblysoi5lq6bj6mhvgiwhy66bk2q",
    )
    base64_Secret_content = (
        get_secret_bundle_by_name_response.data.secret_bundle_content.content
    )
    base64_secret_bytes = base64_Secret_content.encode("ascii")
    base64_message_bytes = base64.b64decode(base64_secret_bytes)
    secret_content = base64_message_bytes.decode("ascii")
    secret_content_json = json.loads(secret_content)
    return secret_content_json

def read_input_file(artifact_type, file_path):
    files_list = []
    workbook = openpyxl.load_workbook(file_path)
    sh1 = workbook[artifact_type]
    row = sh1.max_row
    for r in range(2, row + 1):
        if sh1.cell(row=r, column=1).value == "BI Report":
            files_list.append(
                sh1.cell(row=r, column=3).value + "/" + sh1.cell(row=r, column=2).value
            )
        elif (
            sh1.cell(row=r, column=1).value == "Integrations"
            or sh1.cell(row=r, column=1).value == "Package"
            or sh1.cell(row=r, column=1).value == "Lookup"
            or sh1.cell(row=r, column=1).value == "Library"
        ):
            files_list.append(sh1.cell(row=r, column=2).value)
        elif sh1.cell(row=r, column=1).value == "Database":
            files_list.append(sh1.cell(row=r, column=2).value)
    return files_list

if __name__ == "__main__":
    environment = sys.argv[1]
    artifact_type = sys.argv[2]
    client_git_checkout_name = "client_git"
    release = os.environ["Release_Id"]

    file_path = (
        "./{git}/InputFiles/VBS_Input_{envi}_release{release_number}.xlsx".format(
            git=client_git_checkout_name,
            envi=environment.lower(),
            release_number=release,
        )
    )

    input_file = file_path
    files_list = read_input_file(artifact_type, input_file)
    instance = get_instance(os.environ["Target_Instance"])
    url = instance["URL"]

    # Start dictionary with common entries
    dictionary = {
        "url": url,
        "artifact_list": files_list,
        "source_repo": "./{git}/Release_{release}/{environment}/AP/{artifact_folder}".format(
            git=client_git_checkout_name,
            release=release,
            environment=environment,
            artifact_folder=artifact_type,
        ),
    }

    if artifact_type.lower() == "database":
        # Add database password
        dictionary["username"] = instance["USERNAME"]
        dictionary["password"] = instance["PASSWORD"]
        dictionary["walletpassword"] = instance["WALLETPASSWORD"]
    elif artifact_type.lower() == "integrations":
        # Use integration OAuth flow instead of username/password
        dictionary["scopeurl"] = instance["SCOPE_URL"]
        dictionary["tokenurl"] = instance["TOKEN_URL"]
        dictionary["instanceid"] = instance["INSTANCE_ID"]
        dictionary["clientid"] = instance["CLIENT_ID"]
        dictionary["clientsecret"] = instance["CLIENT_SECRET"]
    else:
        # Use username/password from instance by default for others
        dictionary["username"] = instance["USERNAME"]
        dictionary["password"] = instance["PASSWORD"]

    json_object = json.dumps(dictionary)
    with open("input_param.json", "w") as outfile:
        outfile.write(json_object)
